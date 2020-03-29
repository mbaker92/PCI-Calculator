# Author: Matthew Baker
# Date Modified: 3/29/2020 
# Program: PCICalculator
# Description: Calculates the PCI for the samples imported from a Raw Data csv file.
#               - Imports a User selected Raw Data file to a list of ACPRaw objects
#               - Writes the distress values from each object to the PCI_ACP_Calculator Excel File
#               - Opens Excel to recalculate the PCI value for each object.
#               - Get the calculated values from the Excel file and store them with the object.
#               - Write the newly processed object to the output CSV file.

import openpyxl
from tkinter import *
from tkinter import filedialog as fd
from tkinter import ttk
import formulas
import win32com.client
import os
import pathlib
from ACPRaw import ACPRaw
import csv
import time
import threading

RawDataList=list()      # List for imported CSV
RawDataWithPCI=list()   # List for exported CSV
root= Tk()              # GUI window
fileDirect=""           # File directory
fileName=""             # File Name

# Import selected CSV file to RawDataList 
def ImportRawDataCSV(name):
    # Open the user selected CSV file
    with open(name) as csvfile:
        reader = csv.reader(csvfile,delimiter=',')  # Created reader for csvfile with delimiter ,

        # For loop to get each sample as a ACPRaw object and store each object in the RawDataList
        for row in reader:
            if 'StIDSecID' not in row[0]:   # Skip the Header in the file
                 RawDataList.append(ACPRaw(row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[10], row[11], row[12], row[13], row[14], row[15], row[16],
                                           row[17], row[18], row[19], row[20], row[21], row[22], row[23], row[24], row[25], row[26], row[27], row[28], row[29], row[30], row[31], row[32], row[33],
                                           row[34], row[35], row[36], row[37]))


    # Function used to create the Output CSV file with PCI calculations
def OutputCSV():

    # Open up the save file using the imported file's directory and filename with PCI appended to the end. 
    with open(os.path.normpath(os.path.join(fileDirect,fileName +' PCI.csv')),'w',newline='') as csvFileOut:
        writer=csv.writer(csvFileOut)   # Create the file writer

        # Write the Header row to the file
        writer.writerow(['StIDSecID','Sample Number', 'Rater', 'StreetName','Begin Location', 'End Location', 'Sample Length', 'Sample Width','Date','Sample Notes','Photos','QA','Special','Sample Area','Alligator L','Alligator M','Alligator H','Block L','Block M','Block H',
                        'Distortion L','Distortion M','Distortion H','LongTrans L','LongTrans M','LongTrans H','Patch L', 'Patch M', 'Patch H','Raveling L','Raveling M','Raveling H','RuttingDepression L','RuttingDepression M','RuttingDepression H','Weathering L','Weathering M', 'Weathering H','Calculated PCI'])

        # Write each Sample with PCI to the file
        for Sample in RawDataWithPCI:
            writer.writerow([Sample.StIDSecID,Sample.SampleNumber,Sample.Rater,Sample.StreetName,Sample.BegLocation,Sample.EndLocation,Sample.SampleLength,Sample.SampleWidth,Sample.Date,Sample.SampleNotes,Sample.Photos,Sample.QA,Sample.Special,Sample.SampleArea,
                            Sample.AlligatorL,Sample.AlligatorM,Sample.AlligatorH,Sample.BlockL,Sample.BlockM,Sample.BlockH,Sample.DistortionL,Sample.DistortionM,Sample.DistortionH,Sample.LongTransL,Sample.LongTransM,Sample.LongTransH,
                            Sample.PatchL,Sample.PatchM,Sample.PatchH,Sample.RavelingL,Sample.RavelingM,Sample.RavelingH,Sample.RuttingDepressionL,Sample.RuttingDepressionM,Sample.RuttingDepressionH,Sample.WeatheringL,Sample.WeatheringM,Sample.WeatheringH, Sample.CalcPCI])
            
        print("Exported PCI to - " + os.path.normpath(os.path.join(fileDirect,fileName +' PCI.csv')))   #Notify User of where the file is located


    # Write values of sample to the sheet in the PCI_ACP_Calculator
def WriteToFile(iterat,wb):
        wb['CalcMany'].cell(row=iterat+3,column=2).value=RawDataList[iterat].SampleArea
        wb['CalcMany'].cell(row=iterat+3,column=3).value=RawDataList[iterat].AlligatorL
        wb['CalcMany'].cell(row=iterat+3,column=4).value=RawDataList[iterat].AlligatorM
        wb['CalcMany'].cell(row=iterat+3,column=5).value=RawDataList[iterat].AlligatorH
        wb['CalcMany'].cell(row=iterat+3,column=6).value=RawDataList[iterat].BlockL
        wb['CalcMany'].cell(row=iterat+3,column=7).value=RawDataList[iterat].BlockM
        wb['CalcMany'].cell(row=iterat+3,column=8).value=RawDataList[iterat].BlockH
        wb['CalcMany'].cell(row=iterat+3,column=9).value=RawDataList[iterat].DistortionL
        wb['CalcMany'].cell(row=iterat+3,column=10).value=RawDataList[iterat].DistortionM
        wb['CalcMany'].cell(row=iterat+3,column=11).value=RawDataList[iterat].DistortionH
        wb['CalcMany'].cell(row=iterat+3,column=12).value=RawDataList[iterat].LongTransL
        wb['CalcMany'].cell(row=iterat+3,column=13).value=RawDataList[iterat].LongTransM
        wb['CalcMany'].cell(row=iterat+3,column=14).value=RawDataList[iterat].LongTransH
        wb['CalcMany'].cell(row=iterat+3,column=15).value=RawDataList[iterat].PatchL
        wb['CalcMany'].cell(row=iterat+3,column=16).value=RawDataList[iterat].PatchM
        wb['CalcMany'].cell(row=iterat+3,column=17).value=RawDataList[iterat].PatchH
        wb['CalcMany'].cell(row=iterat+3,column=18).value=RawDataList[iterat].RavelingL
        wb['CalcMany'].cell(row=iterat+3,column=19).value=RawDataList[iterat].RavelingM
        wb['CalcMany'].cell(row=iterat+3,column=20).value=RawDataList[iterat].RavelingH
        wb['CalcMany'].cell(row=iterat+3,column=21).value=RawDataList[iterat].RuttingDepressionL
        wb['CalcMany'].cell(row=iterat+3,column=22).value=RawDataList[iterat].RuttingDepressionM
        wb['CalcMany'].cell(row=iterat+3,column=23).value=RawDataList[iterat].RuttingDepressionH
        wb['CalcMany'].cell(row=iterat+3,column=24).value=RawDataList[iterat].WeatheringL
        wb['CalcMany'].cell(row=iterat+3,column=25).value=RawDataList[iterat].WeatheringM
        wb['CalcMany'].cell(row=iterat+3,column=26).value=RawDataList[iterat].WeatheringH


    #   Open up PCI_ACP_Calculator in Excel to force Excel to recalculate the values
def RefreshExcel(wb):
        print("Calculating Values") # Notify user that the values are being calculated
        wb.save('PCI_ACP_Calculator.xlsx')  # Save the PCI_ACP_Calculator opened in openpyxl

        xlapp = win32com.client.DispatchEx("Excel.Application")     # Create instance of Excel
        file = os.path.abspath('PCI_ACP_Calculator.xlsx')           # Get location of PCI_ACP_Calculator file
        wb2 = xlapp.workbooks.Open(file)                            # Open the file in Excel
        wb2.RefreshAll()                        # Force refresh of the calculations
        wb2.Save()                              # Save the file
        xlapp.Quit()                            # Close Excel


    # Calculate the values using the PCI_ACP_Calculator Excel file
def CalculateValues():
    wb=openpyxl.load_workbook('PCI_ACP_Calculator.xlsx')    # Load the PCI_ACP_Calculator file in openpyxl
     
    counter =0  # Counter for the string notifying user how many rows of the csv are imported into the PCI_ACP_Calculator
    StartingCount=len(RawDataList)  # Get the total number of samples from the csv

    # While there are samples
    while len(RawDataList) > 0:

        if len(RawDataList)>90:     # Use this snippit of code if the len of the list is greater than 90. Only calculating values 90 or less at a time due to limits of Excel Calculator
            for i in range(90):     # Iterate through the first 90 in the list
                WriteToFile(i,wb)   # Write the values to the PCI_ACP_Calculator Excel file
                counter+=1          # Increment the counter displayed to the user
                print('Imported Row ' + str(counter) + " of " +str(StartingCount))  # Notify user of progress

            RefreshExcel(wb)    # Open up the PCI_ACP_Calculator in order to calculated the values we wrote to the file

            wb3=openpyxl.load_workbook('PCI_ACP_Calculator.xlsx',data_only=True)    # Open the PCI_ACP_Calculator with data_only in order to get the calculated value.

            for t in range(90): # Iterate through the first 90 samples in the list again
                RawDataList[t].CalcPCI=wb3['CalcMany'].cell(row=t+3,column=27).value    # Store the Calculated value in the object
                RawDataWithPCI.append(RawDataList[t])   # Copy the object to another list that we will use for exporting the data.
        
            for y in range(90): # Iterate through the first 90 samples again.
                RawDataList.pop(0)  # Remove them from the imported list so the next 90 samples can be processed.

        elif len(RawDataList)<90:           # Use this snippit of code if the len of the list is less than 90
            for last in range(len(RawDataList)):    # Iterate through the list
                WriteToFile(last,wb)        # Write values to the PCI_ACP_Calculator
                counter+=1                  # Increment the counter displayed to the user
                print('Imported Row ' + str(counter) + " of " +str(StartingCount))  # Notify user of progress

            RefreshExcel(wb)   # Open up the PCI_ACP_Calculator in order to calculated the values we wrote to the file

            wb3=openpyxl.load_workbook('PCI_ACP_Calculator.xlsx',data_only=True)    # Open the PCI_ACP_Calculator with data_only in order to get the calculated value.
            for lastt in range(len(RawDataList)): # Iterate through the list again
                RawDataList[lastt].CalcPCI=wb3['CalcMany'].cell(row=lastt+3,column=27).value    # Store the calculated value in the object
                RawDataWithPCI.append(RawDataList[lastt]) # Copy the object to another list that we will use for exporting the data.
            
            RawDataList.clear() # Clear the list
        
        OutputCSV()             # Create the Output CSV
        RawDataWithPCI.clear()  # Clear the Output list for another round of processing.


    # Function for Button Press that starts the whole thing.
def callback():
    # Get the root filename
    root.filename = fd.askopenfilename(initialdir="C:",title="Select Raw Data CSV", filetypes=(("csv files","*.csv"),))
    
    global fileDirect, fileName     # Reference the Global Variables

    # Only run this code if there is a file selected
    if root.filename:       
        ImportRawDataCSV(root.filename)  # Import the Raw Data to the class structure

        print("Importing CSV File") # Notify User the CSV file is importing
        
        # Manipulate the file path of the file selected.
        fileDirect = os.path.dirname(root.filename)     # Get the Directory of the File selected and store it in the Global Variable
        fileName=os.path.basename(root.filename)        # Get the File name with Extension  
        fileName = fileName.split(".")[0]               # Remove the Extension and store it in the Global Variable

        thread1=threading.Thread(target=CalculateValues, args=())   # Create the thread calling the CalculateValues function
        progress(thread1)   # Start the thread


    # Function used for the Progress Bar on the GUI. Will continue showing processing until completed. 
def progress(thread):
    thread.start()    
    pb1= ttk.Progressbar(root,orient=HORIZONTAL, mode='indeterminate') # Progress bar that will continually cycle during processing.
    pb2= ttk.Progressbar(root,orient=HORIZONTAL, mode='determinate')   # Progress bar that will show as completed after processing
    pb2['value']=100    # Set the completed progress bar to 100.

    pb1.pack()  # Add Progress Bar 1 to the GUI
    pb1.start() # Start the Animation

    # Continue the Animation until processing is completed.
    while thread.is_alive():
        root.update()
        pass

    # Destroy Progress Bar 1 and add Progress Bar 2 to the GUI
    pb1.destroy()
    pb2.pack()


root.title('PCI Calculator')    # Window Title
root.minsize(300,100)           # Window Size
button= Button(root,text='Browse',width='30',command=callback)  # Button object for GUI
button.pack()   # Add Button to GUI
root.mainloop() # Start GUI
